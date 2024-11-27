import pandas as pd
import numpy as np
from PIL import Image, ImageDraw, ImageFont
import tkinter as tk
from tkinter import filedialog
import os
from openpyxl import load_workbook

def get_text_width(text, font):
    # 获取文本的实际宽度
    if hasattr(font, 'getlength'):  # PIL 8.0.0 及以上版本
        return font.getlength(str(text))
    else:  # 较老版本的 PIL
        return font.getsize(str(text))[0]

def get_merged_cells(excel_path):
    # 获取合并单元格信息
    wb = load_workbook(excel_path)
    sheet = wb.active
    merged_cells = []
    
    # 获取所有合并单元格的范围
    for merged_range in sheet.merged_cells.ranges:
        min_col = merged_range.min_col - 1  # 转换为0基索引
        max_col = merged_range.max_col - 1
        min_row = merged_range.min_row - 1
        max_row = merged_range.max_row - 1
        merged_cells.append({
            'min_col': min_col,
            'max_col': max_col,
            'min_row': min_row,
            'max_row': max_row
        })
    return merged_cells

def get_merged_cell_value(row, col, merged_cells, df):
    # 获取合并单元格的值（总是使用左上角单元格的值）
    for merged in merged_cells:
        if (merged['min_row'] <= row <= merged['max_row'] and 
            merged['min_col'] <= col <= merged['max_col']):
            # 返回合并区域左上角单元格的值
            return df.iloc[merged['min_row'], merged['min_col']]
    return None

def is_merged_cell(row, col, merged_cells):
    # 检查单元格是否在合并单元格范围内
    for merged in merged_cells:
        if (merged['min_row'] <= row <= merged['max_row'] and 
            merged['min_col'] <= col <= merged['max_col']):
            return merged
    return None

def excel_to_image(excel_path):
    # 读取Excel文件
    df = pd.read_excel(excel_path)
    merged_cells = get_merged_cells(excel_path)
    
    # 设置字体和单元格样式
    font_size = 14
    cell_padding = 10
    header_height = 40
    row_height = 30
    
    try:
        # 尝试加载微软雅黑字体
        font = ImageFont.truetype("msyh.ttc", font_size)
    except:
        # 如果找不到，使用默认字体
        font = ImageFont.load_default()
    
    # 获取所有列的最大宽度，考虑合并单元格的情况
    max_widths = [0] * len(df.columns)
    
    # 计算表头宽度
    for i, col in enumerate(df.columns):
        max_widths[i] = max(max_widths[i], get_text_width(str(col), font))
    
    # 计算数据区域宽度
    for row_idx in range(len(df)):
        for col_idx in range(len(df.columns)):
            value = df.iloc[row_idx, col_idx]
            merged = is_merged_cell(row_idx + 1, col_idx, merged_cells)
            
            if merged:
                if merged['min_col'] == col_idx and merged['min_row'] == row_idx + 1:
                    # 对于合并单元格，计算平均宽度
                    merged_text = str(value) if not pd.isna(value) else ''
                    text_width = get_text_width(merged_text, font)
                    # 将合并单元格的宽度平均分配到每列
                    cols_span = merged['max_col'] - merged['min_col'] + 1
                    avg_width = text_width / cols_span
                    # 更新每列的最大宽度
                    for i in range(merged['min_col'], merged['max_col'] + 1):
                        max_widths[i] = max(max_widths[i], avg_width)
            else:
                if not pd.isna(value):
                    text_width = get_text_width(str(value), font)
                    max_widths[col_idx] = max(max_widths[col_idx], text_width)
    
    # 添加padding
    max_widths = [width + (cell_padding * 2) for width in max_widths]
    
    # 计算图片尺寸
    total_width = int(sum(max_widths))
    total_height = header_height + (len(df) * row_height)
    
    # 创建图片
    image = Image.new('RGB', (total_width, total_height), 'white')
    draw = ImageDraw.Draw(image)
    
    # 绘制表头
    x_pos = 0
    for i, col in enumerate(df.columns):
        merged = is_merged_cell(0, i, merged_cells)
        if merged:
            if merged['min_col'] == i:  # 只在合并单元格的第一个位置绘制
                # 计算合并单元格的总宽度
                merged_width = sum(max_widths[merged['min_col']:merged['max_col'] + 1])
                # 绘制合并单元格背景和边框
                draw.rectangle([
                    (x_pos, 0),
                    (x_pos + merged_width, header_height)
                ], fill='#f0f0f0', outline='#d0d0d0')
                # 绘制文字
                draw.text((x_pos + cell_padding, (header_height - font_size) // 2),
                         str(col), font=font, fill='black')
        else:
            # 绘制普通表头单元格
            draw.rectangle([(x_pos, 0), (x_pos + max_widths[i], header_height)], 
                          fill='#f0f0f0', outline='#d0d0d0')
            draw.text((x_pos + cell_padding, (header_height - font_size) // 2),
                     str(col), font=font, fill='black')
        x_pos += max_widths[i]
    
    # 绘制数据行
    for row_idx, row in df.iterrows():
        y_pos = header_height + row_idx * row_height
        x_pos = 0
        
        for col_idx, value in enumerate(row):
            merged = is_merged_cell(row_idx + 1, col_idx, merged_cells)
            if merged:
                if (merged['min_col'] == col_idx and 
                    merged['min_row'] == row_idx + 1):  # 只在合并单元格的第一个位置绘制
                    # 计算合并单元格的总宽度和高度
                    merged_width = sum(max_widths[merged['min_col']:merged['max_col'] + 1])
                    merged_height = (merged['max_row'] - merged['min_row'] + 1) * row_height
                    # 绘制合并单元格边框
                    draw.rectangle([
                        (x_pos, y_pos),
                        (x_pos + merged_width, y_pos + merged_height)
                    ], outline='#d0d0d0')
                    # 绘制文字
                    cell_text = '' if pd.isna(value) else str(value)
                    draw.text((x_pos + cell_padding, y_pos + (row_height - font_size) // 2),
                             cell_text, font=font, fill='black')
            elif not is_merged_cell(row_idx + 1, col_idx, merged_cells):
                # 只绘制非合并单元格区域内的普通单元格
                draw.rectangle([(x_pos, y_pos), 
                              (x_pos + max_widths[col_idx], y_pos + row_height)],
                             outline='#d0d0d0')
                cell_text = '' if pd.isna(value) else str(value)
                draw.text((x_pos + cell_padding, y_pos + (row_height - font_size) // 2),
                         cell_text, font=font, fill='black')
            x_pos += max_widths[col_idx]
    
    # 保存图片
    output_path = os.path.splitext(excel_path)[0] + '.png'
    image.save(output_path)
    return output_path

def main():
    # 创建tkinter窗口但不显示
    root = tk.Tk()
    root.withdraw()
    
    # 打开文件选择对话框
    excel_path = filedialog.askopenfilename(
        title="选择Excel文件",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    
    if excel_path:
        try:
            output_path = excel_to_image(excel_path)
            print(f"转换成功！图片已保存至: {output_path}")
        except Exception as e:
            print(f"转换过程中出现错误: {str(e)}")
    else:
        print("未选择文件")

if __name__ == "__main__":
    main()
